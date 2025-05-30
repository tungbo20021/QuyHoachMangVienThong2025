import openpyxl
from openpyxl.styles import PatternFill, Font

class NodesExcel:

    def nodes_to_excel(nodes_list, filename="node_info.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Node Info"
        ws.append(["Node", "x", "y", "Traffic"])
        for node in nodes_list:
            ws.append([
                node.get_name(),
                node.get_position_x(),
                node.get_position_y(),
                node.get_traffic()
            ])
        wb.save(filename)
        print(f"Đã lưu file Excel: {filename}")

    def backbones_to_excel(filename, ListMentor):
        try:
            wb = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            print(f"Không tìm thấy file: {filename}")
            return

        ws = wb.create_sheet("Backbones")

        # Định nghĩa màu và font cho header (dòng 1)
        header_fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
        header_font = Font(bold=True)

        for col, group in enumerate(ListMentor, start=1):
            if len(group) == 0:
                continue

            # Ghi số đầu (backbone) vào dòng 1 (header)
            cell = ws.cell(row=1, column=col, value=group[0].get_name())
            cell.fill = header_fill
            cell.font = header_font

            # Ghi các số còn lại từ dòng 2 trở xuống
            for row, node in enumerate(group[1:], start=2):
                ws.cell(row=row, column=col, value=node.get_name())

        wb.save(filename)
        print(f"Đã thêm sheet 'Backbones' vào file: {filename} với header là các backbone.")